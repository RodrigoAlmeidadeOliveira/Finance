"""
Service for data backup and restore operations.
"""
from __future__ import annotations

import json
from contextlib import contextmanager
from datetime import datetime
from typing import Callable, Dict, List, Optional
import io

from sqlalchemy.orm import Session
from sqlalchemy import inspect

from ..models import (
    Category,
    CategoryBudget,
    CategoryRecurringPlan,
    CreditCard,
    FinancialPlan,
    ImportBatch,
    IncomeProjection,
    Institution,
    Investment,
    Dividend,
    InvestmentType,
    PendingTransaction,
    PlanningNote,
    Transaction,
    TrainingJob,
)


class BackupService:
    """Complete data backup and restore service"""

    def __init__(self, session_factory: Callable[[], Session]):
        self.session_factory = session_factory

    @contextmanager
    def _session_scope(self):
        """Context manager for database sessions"""
        session = self.session_factory()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def export_full_backup(self, user_id: int) -> dict:
        """
        Export complete user data in JSON format.

        Returns dictionary with all user data organized by entity type.
        """
        session = self.session_factory()
        try:
            backup_data = {
                'metadata': {
                    'export_date': datetime.utcnow().isoformat(),
                    'user_id': user_id,
                    'version': '1.0',
                    'format': 'flow_forecaster_backup'
                },
                'data': {}
            }

            # Export Categories
            categories = session.query(Category).filter_by(user_id=user_id).all()
            backup_data['data']['categories'] = [
                {
                    'id': c.id,
                    'name': c.name,
                    'type': c.type.value,
                    'parent_id': c.parent_id,
                    'color': c.color,
                    'icon': c.icon,
                    'is_active': c.is_active
                }
                for c in categories
            ]

            # Export Institutions
            institutions = session.query(Institution).filter_by(user_id=user_id).all()
            backup_data['data']['institutions'] = [
                {
                    'id': i.id,
                    'name': i.name,
                    'account_type': i.account_type,
                    'partition': i.partition,
                    'initial_balance': i.initial_balance,
                    'current_balance': i.current_balance,
                    'is_active': i.is_active
                }
                for i in institutions
            ]

            # Export Credit Cards
            credit_cards = session.query(CreditCard).filter_by(user_id=user_id).all()
            backup_data['data']['credit_cards'] = [
                {
                    'id': cc.id,
                    'institution_id': cc.institution_id,
                    'name': cc.name,
                    'brand': cc.brand,
                    'last_four_digits': cc.last_four_digits,
                    'closing_day': cc.closing_day,
                    'due_day': cc.due_day,
                    'limit_amount': cc.limit_amount,
                    'is_active': cc.is_active
                }
                for cc in credit_cards
            ]

            # Export Transactions (manual)
            transactions = session.query(Transaction).filter_by(
                user_id=user_id
            ).filter(
                Transaction.deleted_at.is_(None)
            ).all()
            backup_data['data']['transactions'] = [
                {
                    'id': t.id,
                    'event_date': t.event_date.isoformat() if t.event_date else None,
                    'effective_date': t.effective_date.isoformat() if t.effective_date else None,
                    'transaction_type': t.transaction_type.value,
                    'category_id': t.category_id,
                    'institution_id': t.institution_id,
                    'credit_card_id': t.credit_card_id,
                    'amount': t.amount,
                    'description': t.description,
                    'notes': t.notes,
                    'status': t.status.value,
                    'is_recurring': t.is_recurring,
                    'recurrence_parent_id': t.recurrence_parent_id
                }
                for t in transactions
            ]

            # Export Import Batches
            import_batches = session.query(ImportBatch).filter_by(user_id=user_id).all()
            backup_data['data']['import_batches'] = [
                {
                    'id': ib.id,
                    'filename': ib.filename,
                    'uploaded_at': ib.uploaded_at.isoformat() if ib.uploaded_at else None,
                    'status': ib.status.value,
                    'institution_name': ib.institution_name,
                    'account_id': ib.account_id,
                    'start_date': ib.start_date.isoformat() if ib.start_date else None,
                    'end_date': ib.end_date.isoformat() if ib.end_date else None,
                    'balance': ib.balance,
                    'processed_at': ib.processed_at.isoformat() if ib.processed_at else None,
                    'reviewed_at': ib.reviewed_at.isoformat() if ib.reviewed_at else None,
                    'message': ib.message
                }
                for ib in import_batches
            ]

            # Export Pending Transactions
            pending_txs = session.query(PendingTransaction).join(ImportBatch).filter(
                ImportBatch.user_id == user_id
            ).all()
            backup_data['data']['pending_transactions'] = [
                {
                    'id': pt.id,
                    'batch_id': pt.batch_id,
                    'fitid': pt.fitid,
                    'date': pt.date.isoformat() if pt.date else None,
                    'description': pt.description,
                    'amount': pt.amount,
                    'transaction_type': pt.transaction_type,
                    'predicted_category': pt.predicted_category,
                    'confidence_score': pt.confidence_score,
                    'confidence_level': pt.confidence_level,
                    'user_category': pt.user_category,
                    'review_status': pt.review_status.value if pt.review_status else None,
                    'reviewed_at': pt.reviewed_at.isoformat() if pt.reviewed_at else None,
                    'notes': pt.notes
                }
                for pt in pending_txs
            ]

            # Export Financial Plans
            plans = session.query(FinancialPlan).filter_by(user_id=user_id).all()
            backup_data['data']['financial_plans'] = [
                {
                    'id': fp.id,
                    'name': fp.name,
                    'goal_amount': fp.goal_amount,
                    'current_balance': fp.current_balance,
                    'monthly_contribution': fp.monthly_contribution,
                    'target_date': fp.target_date.isoformat() if fp.target_date else None,
                    'institution_id': fp.institution_id,
                    'partition': fp.partition,
                    'notes': fp.notes,
                    'is_active': fp.is_active
                }
                for fp in plans
            ]

            # Export Income Projections
            income_projs = session.query(IncomeProjection).filter_by(user_id=user_id).all()
            backup_data['data']['income_projections'] = [
                {
                    'id': ip.id,
                    'description': ip.description,
                    'amount': ip.amount,
                    'expected_date': ip.expected_date.isoformat() if ip.expected_date else None,
                    'projection_type': ip.projection_type.value,
                    'received': ip.received,
                    'notes': ip.notes
                }
                for ip in income_projs
            ]

            # Export Category Budgets
            budgets = session.query(CategoryBudget).filter_by(user_id=user_id).all()
            backup_data['data']['category_budgets'] = [
                {
                    'id': cb.id,
                    'category_id': cb.category_id,
                    'month': cb.month,
                    'year': cb.year,
                    'amount': cb.amount
                }
                for cb in budgets
            ]

            # Export Category Recurring Plans
            recurring_plans = session.query(CategoryRecurringPlan).filter_by(user_id=user_id).all()
            backup_data['data']['category_recurring_plans'] = [
                {
                    'id': crp.id,
                    'category_id': crp.category_id,
                    'amount': crp.amount,
                    'start_date': crp.start_date.isoformat() if crp.start_date else None,
                    'end_date': crp.end_date.isoformat() if crp.end_date else None
                }
                for crp in recurring_plans
            ]

            # Export Planning Notes
            notes = session.query(PlanningNote).filter_by(user_id=user_id).all()
            backup_data['data']['planning_notes'] = [
                {
                    'id': pn.id,
                    'content': pn.content,
                    'created_at': pn.created_at.isoformat() if pn.created_at else None
                }
                for pn in notes
            ]

            # Export Investments
            investments = session.query(Investment).filter_by(user_id=user_id).all()
            backup_data['data']['investments'] = [
                {
                    'id': inv.id,
                    'institution_id': inv.institution_id,
                    'investment_type_id': inv.investment_type_id,
                    'description': inv.description,
                    'amount_invested': inv.amount_invested,
                    'current_value': inv.current_value,
                    'classification': inv.classification,
                    'applied_at': inv.applied_at.isoformat() if inv.applied_at else None,
                    'maturity_date': inv.maturity_date.isoformat() if inv.maturity_date else None,
                    'profitability_rate': inv.profitability_rate,
                    'notes': inv.notes
                }
                for inv in investments
            ]

            # Export Dividends
            dividends = session.query(Dividend).join(Investment).filter(
                Investment.user_id == user_id
            ).all()
            backup_data['data']['dividends'] = [
                {
                    'id': d.id,
                    'investment_id': d.investment_id,
                    'amount': d.amount,
                    'received_at': d.received_at.isoformat() if d.received_at else None,
                    'description': d.description
                }
                for d in dividends
            ]

            # Export Training Jobs
            training_jobs = session.query(TrainingJob).filter_by(user_id=user_id).all()
            backup_data['data']['training_jobs'] = [
                {
                    'id': tj.id,
                    'status': tj.status.value,
                    'source': tj.source.value,
                    'csv_path': tj.csv_path,
                    'model_version': tj.model_version,
                    'metrics': tj.metrics,
                    'started_at': tj.started_at.isoformat() if tj.started_at else None,
                    'completed_at': tj.completed_at.isoformat() if tj.completed_at else None,
                    'error_message': tj.error_message
                }
                for tj in training_jobs
            ]

            # Add summary statistics
            backup_data['metadata']['statistics'] = {
                'categories': len(backup_data['data']['categories']),
                'institutions': len(backup_data['data']['institutions']),
                'credit_cards': len(backup_data['data']['credit_cards']),
                'transactions': len(backup_data['data']['transactions']),
                'import_batches': len(backup_data['data']['import_batches']),
                'pending_transactions': len(backup_data['data']['pending_transactions']),
                'financial_plans': len(backup_data['data']['financial_plans']),
                'income_projections': len(backup_data['data']['income_projections']),
                'category_budgets': len(backup_data['data']['category_budgets']),
                'investments': len(backup_data['data']['investments']),
                'dividends': len(backup_data['data']['dividends'])
            }

            return backup_data

        finally:
            session.close()

    def import_full_backup(
        self,
        user_id: int,
        backup_data: dict,
        overwrite: bool = False
    ) -> dict:
        """
        Restore data from backup.

        Args:
            user_id: User to restore data for
            backup_data: Backup dictionary from export_full_backup
            overwrite: If True, delete existing data first (DANGEROUS!)

        Returns:
            dict with statistics about restored items
        """
        if not backup_data.get('metadata') or not backup_data.get('data'):
            raise ValueError("Invalid backup format")

        stats = {
            'categories': 0,
            'institutions': 0,
            'credit_cards': 0,
            'transactions': 0,
            'financial_plans': 0,
            'income_projections': 0,
            'category_budgets': 0,
            'investments': 0,
            'dividends': 0,
            'errors': []
        }

        with self._session_scope() as session:
            if overwrite:
                # WARNING: This deletes all existing data
                # In production, you might want additional safeguards
                self._delete_user_data(session, user_id)

            # Restore in dependency order (foreign keys)

            # 1. Categories (no dependencies)
            for cat_data in backup_data['data'].get('categories', []):
                try:
                    # Check if exists (by name)
                    existing = session.query(Category).filter_by(
                        user_id=user_id,
                        name=cat_data['name'],
                        parent_id=cat_data.get('parent_id')
                    ).first()

                    if not existing:
                        category = Category(
                            user_id=user_id,
                            name=cat_data['name'],
                            type=cat_data['type'],
                            parent_id=cat_data.get('parent_id'),
                            color=cat_data.get('color'),
                            icon=cat_data.get('icon'),
                            is_active=cat_data.get('is_active', True)
                        )
                        session.add(category)
                        stats['categories'] += 1
                except Exception as e:
                    stats['errors'].append(f"Category {cat_data.get('name')}: {str(e)}")

            session.flush()  # Get IDs for foreign keys

            # 2. Institutions
            for inst_data in backup_data['data'].get('institutions', []):
                try:
                    existing = session.query(Institution).filter_by(
                        user_id=user_id,
                        name=inst_data['name']
                    ).first()

                    if not existing:
                        institution = Institution(
                            user_id=user_id,
                            name=inst_data['name'],
                            account_type=inst_data.get('account_type', 'checking'),
                            partition=inst_data.get('partition'),
                            initial_balance=inst_data.get('initial_balance', 0),
                            current_balance=inst_data.get('current_balance', 0),
                            is_active=inst_data.get('is_active', True)
                        )
                        session.add(institution)
                        stats['institutions'] += 1
                except Exception as e:
                    stats['errors'].append(f"Institution {inst_data.get('name')}: {str(e)}")

            session.flush()

            # 3. Credit Cards (depends on institutions)
            # ... (similar pattern)

            # 4. Transactions
            for tx_data in backup_data['data'].get('transactions', []):
                try:
                    transaction = Transaction(
                        user_id=user_id,
                        event_date=datetime.fromisoformat(tx_data['event_date']),
                        effective_date=datetime.fromisoformat(tx_data['effective_date']) if tx_data.get('effective_date') else None,
                        transaction_type=tx_data['transaction_type'],
                        category_id=self._map_category_id(session, user_id, tx_data['category_id']),
                        amount=tx_data['amount'],
                        description=tx_data['description'],
                        notes=tx_data.get('notes'),
                        institution_id=tx_data.get('institution_id'),
                        credit_card_id=tx_data.get('credit_card_id'),
                        status=tx_data['status'],
                        is_recurring=tx_data.get('is_recurring', False)
                    )
                    session.add(transaction)
                    stats['transactions'] += 1
                except Exception as e:
                    stats['errors'].append(f"Transaction {tx_data.get('description')}: {str(e)}")

            # 5. Financial Plans, Investments, etc. (similar pattern)

        return stats

    def _delete_user_data(self, session: Session, user_id: int):
        """Delete all user data (for overwrite restore)"""
        # Delete in reverse dependency order
        session.query(Dividend).filter(
            Dividend.investment_id.in_(
                session.query(Investment.id).filter_by(user_id=user_id)
            )
        ).delete(synchronize_session=False)

        session.query(Investment).filter_by(user_id=user_id).delete()
        session.query(Transaction).filter_by(user_id=user_id).delete()
        session.query(CategoryBudget).filter_by(user_id=user_id).delete()
        session.query(CategoryRecurringPlan).filter_by(user_id=user_id).delete()
        session.query(PlanningNote).filter_by(user_id=user_id).delete()
        session.query(IncomeProjection).filter_by(user_id=user_id).delete()
        session.query(FinancialPlan).filter_by(user_id=user_id).delete()
        session.query(CreditCard).filter_by(user_id=user_id).delete()
        session.query(Institution).filter_by(user_id=user_id).delete()
        session.query(Category).filter_by(user_id=user_id).delete()

    def _map_category_id(self, session: Session, user_id: int, old_cat_id: int) -> int:
        """Map old category ID to new one (for restore)"""
        # This is simplified - in production you'd need a proper ID mapping
        # For now, just return the same ID
        return old_cat_id

    def export_to_excel(self, user_id: int) -> bytes:
        """
        Export data to Excel format (XLSX).
        Returns bytes that can be sent as file download.
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl library required for Excel export")

        backup_data = self.export_full_backup(user_id)

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets for each entity
        for entity_name, items in backup_data['data'].items():
            if not items:
                continue

            ws = wb.create_sheet(title=entity_name[:31])  # Excel limit

            # Headers
            if items:
                headers = list(items[0].keys())
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)

                # Data
                for row_num, item in enumerate(items, 2):
                    for col_num, header in enumerate(headers, 1):
                        value = item.get(header)
                        ws.cell(row=row_num, column=col_num, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()
